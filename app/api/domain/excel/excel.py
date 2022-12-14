import openpyxl as px
from .Model.post import PostModel
from typing import List
import math
import subprocess
import calendar
import os
import shutil
from fastapi.responses import FileResponse

class Excel:
  def __init__(self, post_data: PostModel):
    # 送られてくるデータが昇順であることを前提
    self.ptj_list: List[PtjModel] = post_data.ptj_list
    self.d_week = {'Sun': '日', 'Mon': '月', 'Tue': '火', 'Wed': '水','Thu': '木', 'Fri': '金', 'Sat': '土'}
    # 単一情報
    self.student_number = post_data.user.student_number
    self.name = post_data.user.name
    self.user_type = post_data.user.user_type
    self.subject = post_data.subject.name
    self.teacher_name = post_data.subject.teacher_name

    yaer = self.ptj_list[0].date.strftime('%-Y')
    month = self.ptj_list[0].date.strftime('%-m')
    month_range = calendar.monthrange(int(yaer), int(month))[1]
    self.start_date = f"{yaer}年{month}月1日"
    self.end_date = f"{yaer}年{month}月{month_range}日"
    # ファイル名
    self.file_name = f"{self.student_number}-{self.ptj_list[0].date.strftime('%Y-%-m')}-{self.subject}"
    self.path = f"app/api/domain/excel/files/{self.file_name}"
    self.storage_path = f"{yaer}/{month}/{self.subject}/{self.file_name}.pdf"
    
    if self.user_type == "SA":
      self.wb = px.load_workbook("/app/api/domain/excel/SA.xlsx")
    else:
      self.wb = px.load_workbook("/app/api/domain/excel/TA.xlsx")

  def edit(self):
    ws = self.wb.active
    # SA/TA
    if self.user_type is "TA":
      ws["A1"].value = "ステューデントアシスタント(TA)実施報告書"
    # 学籍番号
    ws["J4"].value = self.student_number
    # 氏名
    ws["X4"].value = "{:　<11}印".format(self.name)
    # 担当科目
    ws["J6"].value = self.subject
    # 担当教員
    ws["J8"].value = self.teacher_name
    # 活動期間
    ws["J10"].value = f"{self.start_date}　　　～　　　　{self.end_date}"
    sum_working_hours = 0.0
    for index, ptj in enumerate(self.ptj_list):
      # 日付
      ws[f"B{14+index}"].value = ptj.date.strftime("%-m月%-d日")
      # 曜日
      ws[f"E{14+index}"].value = self.d_week[ptj.date.strftime("%a")]
      # 開始勤務時間
      # d: 少数部 i: 整数部
      ws[f"H{14+index}"].value = ptj.start_time.strftime("%-H:%M")
      # 終了勤務時間
      ws[f"L{14+index}"].value = ptj.finish_time.strftime("%-H:%M")
      # 合計勤務時間
      ws[f"O{14+index}"].value = f"{ptj.office_hours}時間"
      # 活動内容詳細
      ws[f"S{14+index}"].value = f"{ptj.duties} 休憩:{int(ptj.break_time_minutes)}分"
      sum_working_hours += ptj.office_hours
    # 活動時間計
    ws["O34"].value = f"{sum_working_hours}時間"
    self.wb.save(f"/app/api/domain/excel/files/{self.file_name}.xlsx")

  def convertExcelToPdf(self):
    cmd = []
    cmd.append("libreoffice")
    cmd.append("--headless")
    cmd.append("--nologo")
    cmd.append("--nofirststartwizard")
    cmd.append("--convert-to")
    cmd.append("pdf")
    cmd.append("--outdir")
    cmd.append("/app/api/domain/excel/files")
    cmd.append(f"/app/api/domain/excel/files/{self.file_name}.xlsx")

    subprocess.run(" ".join(cmd), shell=True)

  def pdfCompress(self):
    cmd = []
    cmd.append("-sDEVICE=pdfwrite")
    cmd.append("-dPDFSETTINGS=/printer")
    cmd.append("-dBATCH")
    cmd.append("-dNOPAUSE")
    cmd.append("-dSAFER")
    cmd.append(f"-sOUTPUTFILE=/app/api/domain/excel/files/{self.file_name}.pdf")
    cmd.append(f"/app/api/domain/excel/files/{self.file_name}.pdf")

    subprocess.run(" ".join(cmd), shell=True)
    return FileResponse(path=f"/app/api/domain/excel/files/{self.file_name}.pdf", filename=f"{self.file_name}.pdf")

  def removeFiles(self):
    if(os.path.isdir('/app/api/domain/excel/files') == True):
      shutil.rmtree('/app/api/domain/excel/files')
    os.mkdir('/app/api/domain/excel/files')