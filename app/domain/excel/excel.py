import openpyxl as px
from .Model.post import PostModel
from typing import List
import math

class Excel:
  def __init__(self, post_data: PostModel):
    self.ptj_list: List[PtjModel] = post_data.subject_data.ptjList
    self.d_week = {'Sun': '日', 'Mon': '月', 'Tue': '火', 'Wed': '水','Thu': '木', 'Fri': '金', 'Sat': '土'}
    # 単一情報
    self.student_number = post_data.user.student_number
    self.name = post_data.user.user_name
    self.user_type = post_data.user.user_type
    self.subject = post_data.subject_data.subject
    self.teachers = ",".join(post_data.subject_data.teachers)
    # 送られてくるデータが昇順であることを前提 -> ptj_list[0]:月始め ptj_list[-1]:月終わり
    self.start_date = self.ptj_list[0].date.strftime('%Y年%-m月%-d日')
    self.end_date = self.ptj_list[-1].date.strftime('%Y年%-m月%-d日')

  def edit(self):
    wb = px.load_workbook("/app/domain/excel/SA.xlsx")
    ws = wb.active
    # 学籍番号
    ws["J4"].value = self.student_number
    # 氏名
    ws["X4"].value = "{:　<11}印".format(self.name)
    # 担当科目
    ws["J6"].value = self.subject
    # 担当教員
    ws["J8"].value = self.teachers
    # 活動期間
    ws["J10"].value = f"{self.start_date}　　　～　　　　{self.end_date}"
    for index, ptj in enumerate(self.ptj_list):
      # 日付
      ws[f"B{14+index}"].value = ptj.date.strftime("%-m月%-d日")
      # 曜日
      ws[f"E{14+index}"].value = self.d_week[ptj.date.strftime("%a")]
      # 開始勤務時間
      # d: 少数部 i: 整数部
      d,i = math.modf(ptj.working_start)
      ws[f"H${14+index}"].value = f"{int(i)}:{int(d*100)}"
      # 終了勤務時間
      d,i = math.modf(ptj.working_finish)
      ws[f"L${14+index}"].value = f"{int(i)}:{int(d*100)}"
      # 合計勤務時間
      ws[f"O${14+index}"].value = f"{ptj.sum_working_hours}時間"
      # 活動内容詳細
      ws[f"S{14+index}"].value = f"{ptj.working_content} 休憩:{ptj.break_time}時間"
    wb.save("/app/domain/excel/output.xlsx")